# backend/src/routers/admin_vouchers.py
"""
🎟️ Gestion avancée des Vouchers Admin pour BARRY WiFi
CRUD complet, multi-devices, QR, batch création, export PDF/Excel
Système de génération MASSIVE de vouchers
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, timedelta
from typing import Optional, List
from pydantic import BaseModel
import uuid
import qrcode
import base64
from io import BytesIO
import json

from ..db import get_db
from ..deps import admin_required
from ..models import User, Voucher, Subscription
from ..middleware.admin_logs import admin_logger, LogCategory

router = APIRouter(prefix="/admin/vouchers", tags=["Admin Vouchers"])


# ============================================================
# 📦 CONFIGURATIONS PRÉDÉFINIES DES VOUCHERS
# ============================================================
VOUCHER_PRESETS = {
    # 1️⃣ INDIVIDUAL - Vouchers individuels
    "individual": {
        "pass_500": {"label": "Pass 500", "price": 500, "duration_minutes": 30, "max_devices": 1},
        "pass_1000": {"label": "Pass 1000", "price": 1000, "duration_minutes": 60, "max_devices": 1},
        "pass_2000": {"label": "Pass 2000", "price": 2000, "duration_minutes": 120, "max_devices": 1},
        "pass_5000": {"label": "Pass 5000", "price": 5000, "duration_minutes": 360, "max_devices": 1},
        "1_jour": {"label": "1 Jour", "price": 10000, "duration_minutes": 1440, "max_devices": 1},
        "2_jours": {"label": "2 Jours", "price": 18000, "duration_minutes": 2880, "max_devices": 1},
        "3_jours": {"label": "3 Jours", "price": 25000, "duration_minutes": 4320, "max_devices": 2},
    },
    # 2️⃣ SUBSCRIPTION - Abonnements
    "subscription": {
        "semaine": {"label": "Abonnement Semaine", "price": 50000, "duration_minutes": 10080, "max_devices": 2},
        "mois": {"label": "Abonnement Mensuel", "price": 150000, "duration_minutes": 43200, "max_devices": 3},
        "trimestre": {"label": "Abonnement Trimestriel", "price": 400000, "duration_minutes": 129600, "max_devices": 3},
        "annee": {"label": "Abonnement Annuel", "price": 1500000, "duration_minutes": 525600, "max_devices": 5},
    },
    # 3️⃣ BUSINESS - Entreprise
    "business": {
        "10_employes": {"label": "Entreprise 10 Employés", "price": 500000, "duration_minutes": 43200, "max_devices": 10},
        "30_employes": {"label": "Entreprise 30 Employés", "price": 1200000, "duration_minutes": 43200, "max_devices": 30},
        "50_employes": {"label": "Entreprise 50 Employés", "price": 2000000, "duration_minutes": 43200, "max_devices": 50},
        "illimite": {"label": "Entreprise Illimité", "price": 5000000, "duration_minutes": 43200, "max_devices": 999},
    },
}


# ============================================================
# SCHEMAS
# ============================================================
class VoucherCreateRequest(BaseModel):
    type: str = "individual"  # individual, business, enterprise, vip
    duration_minutes: int = 60
    max_devices: int = 1
    quantity: int = 1
    prefix: Optional[str] = None  # Préfixe personnalisé


class VoucherBulkGenerateRequest(BaseModel):
    """Schéma pour génération massive de vouchers"""
    category: str = "individual"  # individual / subscription / business
    type: str = "individual"      # Type spécifique
    duration_minutes: int = 60
    max_devices: int = 1
    quantity: int = 100           # Peut aller jusqu'à 10000
    price: int = 500
    label: str = "Pass 500"
    prefix: Optional[str] = None
    expires_in_days: Optional[int] = None  # Expiration du voucher non utilisé


class VoucherUpdateRequest(BaseModel):
    type: Optional[str] = None
    duration_minutes: Optional[int] = None
    max_devices: Optional[int] = None
    is_used: Optional[bool] = None
    price: Optional[int] = None
    label: Optional[str] = None


class VoucherBatchResponse(BaseModel):
    created: int
    batch_id: str
    vouchers: List[dict]


# ============================================================
# HELPERS
# ============================================================
# Note: admin_required est maintenant importé depuis deps.py


def generate_voucher_code(prefix: str = None, category: str = "individual") -> str:
    """Génère un code voucher unique selon la catégorie"""
    code = str(uuid.uuid4())[:8].upper()
    
    # Préfixe par catégorie
    category_prefixes = {
        "individual": "IND",
        "subscription": "SUB",
        "business": "BIZ"
    }
    
    cat_prefix = category_prefixes.get(category, "BWF")
    
    if prefix:
        return f"{prefix}-{code}"
    return f"BWIFI-{cat_prefix}-{code}"


def generate_qr_base64(data: str, box_size: int = 10) -> str:
    """Génère un QR code en base64"""
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=box_size,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    return base64.b64encode(buffer.getvalue()).decode()


def generate_batch_id() -> str:
    """Génère un ID unique pour un lot de vouchers"""
    return f"BATCH-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{str(uuid.uuid4())[:6].upper()}"


# ============================================================
# 📋 LISTE DES VOUCHERS
# ============================================================
@router.get("/")
def list_vouchers(
    status: Optional[str] = Query(None, description="used, available, all"),
    type: Optional[str] = Query(None, description="individual, business, enterprise"),
    limit: int = 100,
    offset: int = 0,
    db: Session = Depends(get_db),
    user: User = Depends(admin_required)
):
    """Liste tous les vouchers avec filtres"""
    
    query = db.query(Voucher)
    
    if status == "used":
        query = query.filter(Voucher.is_used == True)
    elif status == "available":
        query = query.filter(Voucher.is_used == False)
    
    if type:
        query = query.filter(Voucher.type == type)
    
    total = query.count()
    vouchers = query.order_by(Voucher.created_at.desc()).limit(limit).offset(offset).all()
    
    return {
        "total": total,
        "count": len(vouchers),
        "vouchers": [
            {
                "id": v.id,
                "code": v.code,
                "type": v.type,
                "duration_minutes": v.duration_minutes,
                "max_devices": v.max_devices,
                "is_used": v.is_used,
                "used_by": v.used_by,
                "used_at": str(v.used_at) if v.used_at else None,
                "created_at": str(v.created_at),
                "qr_data": v.qr_data
            }
            for v in vouchers
        ]
    }


# ============================================================
# ➕ CRÉER UN VOUCHER (Simple)
# ============================================================
@router.post("/create")
def create_voucher(
    data: VoucherCreateRequest,
    db: Session = Depends(get_db),
    user: User = Depends(admin_required)
):
    """Crée un ou plusieurs vouchers (méthode simple)"""
    
    # Validation
    if data.quantity < 1 or data.quantity > 100:
        raise HTTPException(400, "Quantité entre 1 et 100")
    
    if data.duration_minutes < 1:
        raise HTTPException(400, "Durée minimum: 1 minute")
    
    valid_types = ["individual", "business", "enterprise", "vip"]
    if data.type not in valid_types:
        raise HTTPException(400, f"Type invalide. Choix: {valid_types}")
    
    # Ajuster max_devices selon le type
    if data.type == "individual" and data.max_devices > 1:
        data.max_devices = 1
    elif data.type == "business" and data.max_devices < 3:
        data.max_devices = 3
    elif data.type == "enterprise" and data.max_devices < 10:
        data.max_devices = 10
    
    created_vouchers = []
    batch_id = generate_batch_id()
    
    for _ in range(data.quantity):
        code = generate_voucher_code(data.prefix, "individual")
        
        # Vérifier unicité
        while db.query(Voucher).filter(Voucher.code == code).first():
            code = generate_voucher_code(data.prefix, "individual")
        
        qr_data = generate_qr_base64(code)
        
        voucher = Voucher(
            code=code,
            category="individual",
            type=data.type,
            duration_minutes=data.duration_minutes,
            max_devices=data.max_devices,
            qr_data=qr_data,
            batch_id=batch_id,
            created_by=user.id,
            created_at=datetime.utcnow()
        )
        
        db.add(voucher)
        created_vouchers.append({
            "code": code,
            "type": data.type,
            "duration_minutes": data.duration_minutes,
            "max_devices": data.max_devices
        })
    
    db.commit()
    
    # Logger
    admin_logger.log_voucher_action(
        action="vouchers_created",
        admin_id=user.id,
        description=f"Création de {data.quantity} voucher(s) de type {data.type}"
    )
    
    return {
        "ok": True,
        "created": len(created_vouchers),
        "batch_id": batch_id,
        "vouchers": created_vouchers
    }


# ============================================================
# 🚀 GÉNÉRATION MASSIVE DE VOUCHERS
# ============================================================
@router.post("/generate-bulk")
def generate_bulk_vouchers(
    data: VoucherBulkGenerateRequest,
    db: Session = Depends(get_db),
    user: User = Depends(admin_required)
):
    """
    🚀 Génère massivement des vouchers (jusqu'à 10000)
    
    Catégories disponibles:
    - individual: Vouchers individuels (500/1000/2000/1 jour/2 jours)
    - subscription: Abonnements (semaine, mois, trimestre, année)
    - business: Entreprise (10/30/50 employés, illimité)
    """
    
    # Validation quantité (jusqu'à 10000)
    if data.quantity < 1:
        raise HTTPException(400, "Quantité minimum: 1")
    if data.quantity > 10000:
        raise HTTPException(400, "Quantité maximum: 10000 vouchers par lot")
    
    # Validation catégorie
    valid_categories = ["individual", "subscription", "business"]
    if data.category not in valid_categories:
        raise HTTPException(400, f"Catégorie invalide. Choix: {valid_categories}")
    
    # Validation durée
    if data.duration_minutes < 1:
        raise HTTPException(400, "Durée minimum: 1 minute")
    
    # Validation prix
    if data.price < 0:
        raise HTTPException(400, "Prix invalide")
    
    # Génération du batch ID
    batch_id = generate_batch_id()
    
    # Calcul expiration si spécifié
    expires_at = None
    if data.expires_in_days:
        expires_at = datetime.utcnow() + timedelta(days=data.expires_in_days)
    
    created_vouchers = []
    
    # Génération optimisée en batch
    for i in range(data.quantity):
        code = generate_voucher_code(data.prefix, data.category)
        
        # Vérifier unicité
        while db.query(Voucher).filter(Voucher.code == code).first():
            code = generate_voucher_code(data.prefix, data.category)
        
        # Générer QR code (optimisé: taille réduite pour grandes quantités)
        qr_box_size = 6 if data.quantity > 500 else 10
        qr_data = generate_qr_base64(code, qr_box_size)
        
        voucher = Voucher(
            code=code,
            category=data.category,
            type=data.type,
            label=data.label,
            duration_minutes=data.duration_minutes,
            max_devices=data.max_devices,
            price=data.price,
            currency="GNF",
            status="unused",
            qr_data=qr_data,
            batch_id=batch_id,
            expires_at=expires_at,
            created_by=user.id,
            created_at=datetime.utcnow()
        )
        
        db.add(voucher)
        
        # Ajouter à la liste (sans QR pour réponse légère)
        created_vouchers.append({
            "id": None,  # Sera assigné après commit
            "code": code,
            "category": data.category,
            "type": data.type,
            "label": data.label,
            "duration_minutes": data.duration_minutes,
            "max_devices": data.max_devices,
            "price": data.price,
            "status": "unused",
            "qr_data": qr_data if data.quantity <= 100 else None,  # QR uniquement pour petits lots
            "created_at": datetime.utcnow().isoformat()
        })
    
    db.commit()
    
    # Logger l'action
    admin_logger.log_voucher_action(
        action="bulk_vouchers_generated",
        admin_id=user.id,
        description=f"Génération massive de {data.quantity} vouchers [{data.category}] - Lot: {batch_id}"
    )
    
    return {
        "ok": True,
        "message": f"✅ {data.quantity} vouchers générés avec succès",
        "batch_id": batch_id,
        "created": len(created_vouchers),
        "category": data.category,
        "label": data.label,
        "total_value": data.price * data.quantity,
        "currency": "GNF",
        "vouchers": created_vouchers
    }


# ============================================================
# 📋 RÉCUPÉRER LES PRESETS DISPONIBLES
# ============================================================
@router.get("/presets")
def get_voucher_presets(
    user: User = Depends(admin_required)
):
    """Retourne les configurations prédéfinies des vouchers"""
    return {
        "presets": VOUCHER_PRESETS,
        "categories": ["individual", "subscription", "business"]
    }


# ============================================================
# 📦 RÉCUPÉRER UN LOT PAR BATCH ID
# ============================================================
@router.get("/batch/{batch_id}")
def get_vouchers_by_batch(
    batch_id: str,
    include_qr: bool = False,
    db: Session = Depends(get_db),
    user: User = Depends(admin_required)
):
    """Récupère tous les vouchers d'un lot"""
    
    vouchers = db.query(Voucher).filter(Voucher.batch_id == batch_id).all()
    
    if not vouchers:
        raise HTTPException(404, "Lot non trouvé")
    
    return {
        "batch_id": batch_id,
        "count": len(vouchers),
        "vouchers": [
            {
                "id": v.id,
                "code": v.code,
                "category": v.category,
                "type": v.type,
                "label": v.label,
                "duration_minutes": v.duration_minutes,
                "max_devices": v.max_devices,
                "price": v.price,
                "status": v.status,
                "is_used": v.is_used,
                "qr_data": v.qr_data if include_qr else None,
                "created_at": str(v.created_at)
            }
            for v in vouchers
        ]
    }


# ============================================================
# 📖 DÉTAIL D'UN VOUCHER
# ============================================================
@router.get("/{voucher_id}")
def get_voucher_detail(
    voucher_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(admin_required)
):
    """Détails complets d'un voucher"""
    
    voucher = db.query(Voucher).filter(Voucher.id == voucher_id).first()
    if not voucher:
        raise HTTPException(404, "Voucher non trouvé")
    
    # Infos utilisateur si utilisé
    used_by_user = None
    if voucher.used_by:
        u = db.query(User).filter(User.id == voucher.used_by).first()
        if u:
            used_by_user = {
                "id": u.id,
                "name": f"{u.first_name} {u.last_name}",
                "phone": u.phone_number
            }
    
    # Sessions actives (pour business/enterprise)
    active_sessions = 0
    if voucher.type in ["business", "enterprise"]:
        active_sessions = db.query(Subscription).filter(
            Subscription.voucher_code == voucher.code,
            Subscription.is_active == True,
            Subscription.end_at > datetime.utcnow()
        ).count()
    
    return {
        "id": voucher.id,
        "code": voucher.code,
        "type": voucher.type,
        "duration_minutes": voucher.duration_minutes,
        "max_devices": voucher.max_devices,
        "is_used": voucher.is_used,
        "used_by": used_by_user,
        "used_at": str(voucher.used_at) if voucher.used_at else None,
        "active_sessions": active_sessions,
        "remaining_devices": voucher.max_devices - active_sessions,
        "created_at": str(voucher.created_at),
        "qr_data": voucher.qr_data
    }


# ============================================================
# ✏️ MODIFIER UN VOUCHER
# ============================================================
@router.put("/{voucher_id}")
def update_voucher(
    voucher_id: int,
    data: VoucherUpdateRequest,
    db: Session = Depends(get_db),
    user: User = Depends(admin_required)
):
    """Modifie un voucher existant"""
    
    voucher = db.query(Voucher).filter(Voucher.id == voucher_id).first()
    if not voucher:
        raise HTTPException(404, "Voucher non trouvé")
    
    if data.type:
        voucher.type = data.type
    if data.duration_minutes:
        voucher.duration_minutes = data.duration_minutes
    if data.max_devices:
        voucher.max_devices = data.max_devices
    if data.is_used is not None:
        voucher.is_used = data.is_used
        if not data.is_used:
            voucher.used_by = None
            voucher.used_at = None
    
    db.commit()
    
    admin_logger.log_voucher_action(
        action="voucher_updated",
        voucher_code=voucher.code,
        admin_id=user.id,
        description=f"Voucher {voucher.code} modifié"
    )
    
    return {"ok": True, "message": "Voucher modifié"}


# ============================================================
# 🗑️ SUPPRIMER UN VOUCHER
# ============================================================
@router.delete("/{voucher_id}")
def delete_voucher(
    voucher_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(admin_required)
):
    """Supprime un voucher (si non utilisé)"""
    
    voucher = db.query(Voucher).filter(Voucher.id == voucher_id).first()
    if not voucher:
        raise HTTPException(404, "Voucher non trouvé")
    
    if voucher.is_used:
        raise HTTPException(400, "Impossible de supprimer un voucher déjà utilisé")
    
    code = voucher.code
    db.delete(voucher)
    db.commit()
    
    admin_logger.log_voucher_action(
        action="voucher_deleted",
        voucher_code=code,
        admin_id=user.id,
        description=f"Voucher {code} supprimé"
    )
    
    return {"ok": True, "message": "Voucher supprimé"}


# ============================================================
# 🚫 DÉSACTIVER UN VOUCHER
# ============================================================
@router.post("/{voucher_id}/disable")
def disable_voucher(
    voucher_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(admin_required)
):
    """Désactive un voucher (le marque comme utilisé)"""
    
    voucher = db.query(Voucher).filter(Voucher.id == voucher_id).first()
    if not voucher:
        raise HTTPException(404, "Voucher non trouvé")
    
    voucher.is_used = True
    voucher.used_at = datetime.utcnow()
    db.commit()
    
    admin_logger.log_voucher_action(
        action="voucher_disabled",
        voucher_code=voucher.code,
        admin_id=user.id,
        description=f"Voucher {voucher.code} désactivé manuellement"
    )
    
    return {"ok": True, "message": "Voucher désactivé"}


# ============================================================
# ✅ RÉACTIVER UN VOUCHER
# ============================================================
@router.post("/{voucher_id}/enable")
def enable_voucher(
    voucher_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(admin_required)
):
    """Réactive un voucher (le marque comme non utilisé)"""
    
    voucher = db.query(Voucher).filter(Voucher.id == voucher_id).first()
    if not voucher:
        raise HTTPException(404, "Voucher non trouvé")
    
    voucher.is_used = False
    voucher.used_by = None
    voucher.used_at = None
    db.commit()
    
    admin_logger.log_voucher_action(
        action="voucher_enabled",
        voucher_code=voucher.code,
        admin_id=user.id,
        description=f"Voucher {voucher.code} réactivé"
    )
    
    return {"ok": True, "message": "Voucher réactivé"}


# ============================================================
# 📊 STATISTIQUES VOUCHERS
# ============================================================
@router.get("/stats/summary")
def get_voucher_stats(
    db: Session = Depends(get_db),
    user: User = Depends(admin_required)
):
    """Statistiques des vouchers"""
    
    total = db.query(func.count(Voucher.id)).scalar() or 0
    used = db.query(func.count(Voucher.id)).filter(Voucher.is_used == True).scalar() or 0
    available = total - used
    
    by_type = db.query(
        Voucher.type,
        func.count(Voucher.id)
    ).group_by(Voucher.type).all()
    
    return {
        "total": total,
        "used": used,
        "available": available,
        "usage_rate": round((used / total * 100), 1) if total > 0 else 0,
        "by_type": {t[0]: t[1] for t in by_type}
    }


# ============================================================
# 🔍 RECHERCHER UN VOUCHER PAR CODE
# ============================================================
@router.get("/search/{code}")
def search_voucher(
    code: str,
    db: Session = Depends(get_db),
    user: User = Depends(admin_required)
):
    """Recherche un voucher par son code"""
    
    voucher = db.query(Voucher).filter(Voucher.code == code.upper()).first()
    if not voucher:
        raise HTTPException(404, "Voucher non trouvé")
    
    return {
        "id": voucher.id,
        "code": voucher.code,
        "category": getattr(voucher, 'category', 'individual'),
        "type": voucher.type,
        "label": getattr(voucher, 'label', None),
        "duration_minutes": voucher.duration_minutes,
        "max_devices": voucher.max_devices,
        "price": getattr(voucher, 'price', 0),
        "status": getattr(voucher, 'status', 'unused' if not voucher.is_used else 'used'),
        "is_used": voucher.is_used,
        "created_at": str(voucher.created_at)
    }


# ============================================================
# 📄 EXPORT PDF DES VOUCHERS
# ============================================================
@router.get("/export/pdf")
def export_vouchers_pdf(
    batch_id: Optional[str] = None,
    status: Optional[str] = None,
    category: Optional[str] = None,
    limit: int = 500,
    db: Session = Depends(get_db),
    user: User = Depends(admin_required)
):
    """
    📄 Exporte les vouchers en PDF avec QR codes
    - batch_id: Filtrer par lot
    - status: used, unused, all
    - category: individual, subscription, business
    """
    
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        raise HTTPException(500, "Module reportlab non installé. Exécutez: pip install reportlab")
    
    # Requête de base
    query = db.query(Voucher)
    
    if batch_id:
        query = query.filter(Voucher.batch_id == batch_id)
    
    if status == "used":
        query = query.filter(Voucher.is_used == True)
    elif status == "unused":
        query = query.filter(Voucher.is_used == False)
    
    if category:
        query = query.filter(Voucher.category == category)
    
    vouchers = query.order_by(Voucher.created_at.desc()).limit(limit).all()
    
    if not vouchers:
        raise HTTPException(404, "Aucun voucher trouvé")
    
    # Créer le PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=20
    )
    
    elements = []
    
    # Titre
    elements.append(Paragraph("🎟️ BARRY WI-FI - Liste des Vouchers", title_style))
    elements.append(Paragraph(f"Date: {datetime.utcnow().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Paragraph(f"Total: {len(vouchers)} vouchers", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Tableau des vouchers
    data = [["#", "Code", "Catégorie", "Label", "Prix", "Durée", "Statut"]]
    
    for i, v in enumerate(vouchers, 1):
        duration_str = f"{v.duration_minutes} min"
        if v.duration_minutes >= 1440:
            duration_str = f"{v.duration_minutes // 1440} jour(s)"
        
        status_text = "✅ Disponible" if not v.is_used else "❌ Utilisé"
        price_str = f"{getattr(v, 'price', 0):,} GNF".replace(',', ' ')
        
        data.append([
            str(i),
            v.code,
            getattr(v, 'category', 'individual').upper(),
            getattr(v, 'label', '-') or '-',
            price_str,
            duration_str,
            status_text
        ])
    
    table = Table(data, colWidths=[30, 120, 70, 100, 80, 60, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
    ]))
    
    elements.append(table)
    
    # Construire le PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"vouchers_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================================
# 📊 EXPORT EXCEL DES VOUCHERS
# ============================================================
@router.get("/export/excel")
def export_vouchers_excel(
    batch_id: Optional[str] = None,
    status: Optional[str] = None,
    category: Optional[str] = None,
    limit: int = 5000,
    db: Session = Depends(get_db),
    user: User = Depends(admin_required)
):
    """
    📊 Exporte les vouchers en Excel
    - batch_id: Filtrer par lot
    - status: used, unused, all
    - category: individual, subscription, business
    """
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "Module openpyxl non installé. Exécutez: pip install openpyxl")
    
    # Requête de base
    query = db.query(Voucher)
    
    if batch_id:
        query = query.filter(Voucher.batch_id == batch_id)
    
    if status == "used":
        query = query.filter(Voucher.is_used == True)
    elif status == "unused":
        query = query.filter(Voucher.is_used == False)
    
    if category:
        query = query.filter(Voucher.category == category)
    
    vouchers = query.order_by(Voucher.created_at.desc()).limit(limit).all()
    
    if not vouchers:
        raise HTTPException(404, "Aucun voucher trouvé")
    
    # Créer le workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Vouchers"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    centered = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes
    headers = ["ID", "Code", "Catégorie", "Type", "Label", "Prix (GNF)", 
               "Durée (min)", "Max Devices", "Statut", "Utilisé", "Batch ID", 
               "Date Création", "Date Utilisation"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
        cell.border = thin_border
    
    # Données
    for row, v in enumerate(vouchers, 2):
        ws.cell(row=row, column=1, value=v.id).border = thin_border
        ws.cell(row=row, column=2, value=v.code).border = thin_border
        ws.cell(row=row, column=3, value=getattr(v, 'category', 'individual')).border = thin_border
        ws.cell(row=row, column=4, value=v.type).border = thin_border
        ws.cell(row=row, column=5, value=getattr(v, 'label', '')).border = thin_border
        ws.cell(row=row, column=6, value=getattr(v, 'price', 0)).border = thin_border
        ws.cell(row=row, column=7, value=v.duration_minutes).border = thin_border
        ws.cell(row=row, column=8, value=v.max_devices).border = thin_border
        ws.cell(row=row, column=9, value=getattr(v, 'status', 'unused' if not v.is_used else 'used')).border = thin_border
        ws.cell(row=row, column=10, value="Oui" if v.is_used else "Non").border = thin_border
        ws.cell(row=row, column=11, value=getattr(v, 'batch_id', '')).border = thin_border
        ws.cell(row=row, column=12, value=str(v.created_at) if v.created_at else '').border = thin_border
        ws.cell(row=row, column=13, value=str(v.used_at) if v.used_at else '').border = thin_border
    
    # Ajuster largeur des colonnes
    column_widths = [8, 25, 12, 12, 25, 15, 12, 12, 12, 10, 30, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Sauvegarder en mémoire
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"vouchers_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================================
# 📊 STATISTIQUES AVANCÉES PAR CATÉGORIE
# ============================================================
@router.get("/stats/by-category")
def get_voucher_stats_by_category(
    db: Session = Depends(get_db),
    user: User = Depends(admin_required)
):
    """Statistiques détaillées des vouchers par catégorie"""
    
    # Stats globales
    total = db.query(func.count(Voucher.id)).scalar() or 0
    used = db.query(func.count(Voucher.id)).filter(Voucher.is_used == True).scalar() or 0
    available = total - used
    
    # Stats par catégorie
    categories_stats = {}
    for category in ["individual", "subscription", "business"]:
        cat_total = db.query(func.count(Voucher.id)).filter(Voucher.category == category).scalar() or 0
        cat_used = db.query(func.count(Voucher.id)).filter(
            Voucher.category == category,
            Voucher.is_used == True
        ).scalar() or 0
        
        # Revenu potentiel et réalisé
        revenue_potential = db.query(func.sum(Voucher.price)).filter(
            Voucher.category == category
        ).scalar() or 0
        
        revenue_realized = db.query(func.sum(Voucher.price)).filter(
            Voucher.category == category,
            Voucher.is_used == True
        ).scalar() or 0
        
        categories_stats[category] = {
            "total": cat_total,
            "used": cat_used,
            "available": cat_total - cat_used,
            "usage_rate": round((cat_used / cat_total * 100), 1) if cat_total > 0 else 0,
            "revenue_potential": revenue_potential,
            "revenue_realized": revenue_realized
        }
    
    # Stats par batch (les 10 derniers)
    recent_batches = db.query(
        Voucher.batch_id,
        func.count(Voucher.id).label('count'),
        func.min(Voucher.created_at).label('created_at')
    ).filter(
        Voucher.batch_id != None
    ).group_by(
        Voucher.batch_id
    ).order_by(
        func.min(Voucher.created_at).desc()
    ).limit(10).all()
    
    return {
        "global": {
            "total": total,
            "used": used,
            "available": available,
            "usage_rate": round((used / total * 100), 1) if total > 0 else 0
        },
        "by_category": categories_stats,
        "recent_batches": [
            {
                "batch_id": b.batch_id,
                "count": b.count,
                "created_at": str(b.created_at) if b.created_at else None
            }
            for b in recent_batches
        ]
    }

